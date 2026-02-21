"""
题目导入器 — JSON / CSV / Excel 批量导入
"""

import json
import os
from pathlib import Path

from .validator import validate_batch, validate_question, ValidationResult


def load_json_file(path: str | Path) -> tuple[dict, list[dict]]:
    """读取 JSON 文件，返回 (batch_meta, questions)"""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")
    if path.suffix.lower() not in (".json",):
        raise ValueError(f"不支持的文件格式: {path.suffix}")

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    batch_meta = {
        "batch_name": data.get("batch_name", path.stem),
        "source": data.get("source", "custom"),
        "source_year": data.get("source_year"),
        "subject": data.get("subject"),
        "grade_level": data.get("grade_level"),
    }

    raw_questions = data.get("questions", [])
    if not raw_questions:
        raise ValueError("JSON 文件中未找到 questions 数组")

    # 将 batch 级别默认值注入每个题目
    questions = []
    for q in raw_questions:
        merged = {
            "subject": batch_meta.get("subject"),
            "grade_level": batch_meta.get("grade_level"),
            "source": batch_meta.get("source"),
            "source_year": batch_meta.get("source_year"),
            "question_type": "mcq",
            "review_status": "draft",
        }
        merged.update(q)
        questions.append(merged)

    return batch_meta, questions


def load_csv_file(path: str | Path) -> tuple[dict, list[dict]]:
    """读取 CSV/Excel 文件，返回 (batch_meta, questions)

    支持 .csv 和 .xlsx。Excel 需要 openpyxl。
    列名映射:
      stem_zh, stem_en, option_a/b/c/d, answer, topic, subtopic,
      difficulty, grade_level, explanation, subject, question_type
    """
    import csv

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {path}")

    rows: list[dict] = []

    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("导入 Excel 需要 openpyxl: pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [v if v is not None else "" for v in row])))
        wb.close()
    elif path.suffix.lower() == ".csv":
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({k.strip().lower(): v for k, v in row.items()})
    else:
        raise ValueError(f"不支持的文件格式: {path.suffix}")

    batch_meta = {
        "batch_name": path.stem,
        "source": "custom",
    }

    questions = []
    for row in rows:
        stem_zh = row.get("stem_zh", "").strip()
        stem_en = row.get("stem_en", "").strip()
        if not stem_zh and not stem_en:
            continue

        options = []
        for key in ("option_a", "option_b", "option_c", "option_d"):
            val = str(row.get(key, "")).strip()
            if val:
                letter = key[-1].upper()
                options.append(f"{letter}. {val}")

        answer = str(row.get("answer", "")).strip().upper()
        q_type = str(row.get("question_type", "mcq")).strip()

        content_zh = {"stem": stem_zh, "options": options, "answer": answer} if options else {"stem": stem_zh, "answer": answer}
        content_en = {"stem": stem_en, "options": options, "answer": answer} if stem_en else content_zh

        difficulty = row.get("difficulty")
        try:
            difficulty = float(difficulty) if difficulty else 0.5
        except (ValueError, TypeError):
            difficulty = 0.5

        questions.append({
            "subject": str(row.get("subject", "math")).strip(),
            "grade_level": str(row.get("grade_level", "G7")).strip(),
            "topic": str(row.get("topic", "general")).strip(),
            "subtopic": str(row.get("subtopic", "")).strip() or None,
            "difficulty": difficulty,
            "question_type": q_type,
            "content_zh": content_zh,
            "content_en": content_en,
            "explanation_zh": str(row.get("explanation", "")).strip() or None,
            "explanation_en": str(row.get("explanation_en", "")).strip() or None,
            "review_status": "draft",
            "source": "custom",
        })

    return batch_meta, questions


async def import_questions(
    questions: list[dict],
    batch_meta: dict,
    *,
    imported_by: str | None = None,
    validate: bool = True,
) -> dict:
    """导入题目到数据库，返回 {batch_id, imported, errors, warnings}"""
    from .. import db

    # 校验
    if validate:
        result = validate_batch(questions)
        if not result.valid:
            return {
                "batch_id": None,
                "imported": 0,
                "errors": result.errors,
                "warnings": result.warnings,
            }

    # 创建导入批次
    batch = await db.create_import_batch(
        batch_name=batch_meta.get("batch_name", "unnamed"),
        source=batch_meta.get("source", "custom"),
        imported_by=imported_by,
    )
    batch_id = batch["id"]

    try:
        count = await db.bulk_insert_questions(questions, batch_id=batch_id)
        await db.update_import_batch(batch_id, status="imported", question_count=count)
        result = validate_batch(questions) if validate else ValidationResult()
        return {
            "batch_id": batch_id,
            "imported": count,
            "errors": [],
            "warnings": result.warnings,
        }
    except Exception as e:
        await db.update_import_batch(batch_id, status="failed", error_log=str(e))
        return {
            "batch_id": batch_id,
            "imported": 0,
            "errors": [str(e)],
            "warnings": [],
        }


def scan_directory(dir_path: str | Path, suffix: str = ".json") -> list[Path]:
    """递归扫描目录中的文件"""
    dir_path = Path(dir_path)
    if not dir_path.is_dir():
        raise NotADirectoryError(f"不是目录: {dir_path}")
    return sorted(dir_path.rglob(f"*{suffix}"))
